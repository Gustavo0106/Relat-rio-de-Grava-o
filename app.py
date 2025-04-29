import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Relatório de Gravação", layout="centered")

uploaded_file = st.file_uploader("📁 Selecione a planilha .xlsx com os dados", type="xlsx")

if uploaded_file:
    # Leitura do Excel
    xls = pd.ExcelFile(uploaded_file)
    df = xls.parse("Planilha1")

    # Converte os dados
    df_numeric = df.drop(columns=['MOOVSEC']).notna().astype(int)
    df_numeric['MOOVSEC'] = df['MOOVSEC']
    cols = ['MOOVSEC'] + [col for col in df_numeric.columns if col != 'MOOVSEC']
    df_numeric = df_numeric[cols]

    # Calcula dias e porcentagem
    df_numeric['Dias Gravados'] = df_numeric.drop(columns=['MOOVSEC']).sum(axis=1)
    total_dias = df_numeric.shape[1] - 2
    df_numeric['% Gravação'] = (df_numeric['Dias Gravados'] / total_dias * 100).round(1)

    # Situação com alerta
    df_numeric['Situação'] = np.where(
        df_numeric['% Gravação'] >= 80, '✅ OK',
        np.where(df_numeric['% Gravação'] >= 50, '⚠️ Média', '❌ Baixa')
    )

    # Resumo
    resumo = df_numeric[['MOOVSEC', 'Dias Gravados', '% Gravação', 'Situação']]

    st.markdown("<h2 style='text-align: center; color: #4CAF50;'>📋 Resumo dos Dados</h2>", unsafe_allow_html=True)
    st.dataframe(resumo, use_container_width=True)

    # Geração do Excel com estilo
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    for r in dataframe_to_rows(resumo, index=False, header=True):
        ws.append(r)

    # Estilo
    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4):
        situacao_cell = row[3]
        if situacao_cell.value == '✅ OK':
            for cell in row:
                cell.fill = fill_verde
        elif situacao_cell.value == '⚠️ Média':
            for cell in row:
                cell.fill = fill_amarelo
        elif situacao_cell.value == '❌ Baixa':
            for cell in row:
                cell.fill = fill_vermelho

    for cell in ws[1]:
        cell.font = bold_font

    # Salva em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Botão de download
    st.success("✅ Relatório gerado com sucesso!")
    st.download_button(
        label="📥 Baixar Relatório Excel",
        data=output,
        file_name="Relatorio_de_Gravacao.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

